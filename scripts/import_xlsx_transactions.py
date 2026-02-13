"""One-time import of Morgan Stanley money movement transactions from XLSX.

Reads a Morgan Stanley "Activity Report" XLSX file and imports transactions
into plaid_transactions + derived_transactions tables.

Data integrity guarantees:
  - Deterministic external_ids via SHA-256 hash of (date, description,
    amount, account_id, sequence). Re-running produces identical IDs.
  - Source = XLSX_IMPORT keeps these distinct from Plaid-synced data.
  - ON CONFLICT (external_id, source) upsert makes the script idempotent.
  - Derived rows are only inserted for plaid_transaction_ids that don't
    already have derived rows, preventing duplicates on re-run.
  - Accounts already covered by Plaid are skipped by default to avoid
    real-world duplicate transactions.

Usage:
    # Dry run (default): show what would be imported
    uv run python scripts/import_xlsx_transactions.py \
        --xlsx "~/Downloads/Activity Report.xlsx"

    # Actually import
    uv run python scripts/import_xlsx_transactions.py \
        --xlsx "~/Downloads/Activity Report.xlsx" \
        --apply

    # Import including accounts covered by Plaid (creates duplicates)
    uv run python scripts/import_xlsx_transactions.py \
        --xlsx "~/Downloads/Activity Report.xlsx" \
        --apply --include-plaid-covered

    # Override item_id (e.g. after reconnecting Morgan Stanley)
    uv run python scripts/import_xlsx_transactions.py \
        --xlsx "~/Downloads/Activity Report.xlsx" \
        --apply --item-id <new-item-id>

Original import: 2026-02-13
  - 531 transactions across 8 Morgan Stanley accounts
  - Skipped 2,563 American Express rows (covered by Plaid)
"""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import date, datetime
import hashlib
import os
from pathlib import Path
import sys

from dotenv import load_dotenv
import openpyxl

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SOURCE = "XLSX_IMPORT"
DEFAULT_ITEM_ID = "mJVQYE870yhdDRZgZzDMHZJYqPK0MaHM3ppm7"

# XLSX account patterns to skip because Plaid already covers them.
# Map of pattern → reason. Use --include-plaid-covered to override.
PLAID_COVERED_PATTERNS: dict[str, str] = {
    "1008": (
        "American Express Gold Card — covered by Plaid (3,837 txns, Jan 2024–present)"
    ),
}

# Header row index (1-based) in the XLSX sheet
HEADER_ROW = 7
DATA_START_ROW = HEADER_ROW + 1

# Expected columns at HEADER_ROW
EXPECTED_HEADERS = [
    "Activity Date",
    "Transaction Date",
    "Account",
    "Institution Name",
    "Activity",
    "Description",
    "Category",
    "Memo",
    "Tags",
    "Amount($)",
]


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------


def extract_account_id(account_name: str) -> str:
    """Extract the trailing identifier from account name.

    Examples:
        '2. Adam - Sep. Prop - Brokerage - 9320' -> '9320'
        'AFFIRM, INC. - 033H' -> '033H'
    """
    parts = account_name.rsplit(" - ", 1)
    if len(parts) == 2:
        return parts[1].strip()
    return account_name.strip()


def resolve_institution(account_name: str) -> str:
    """Resolve institution name from XLSX account name."""
    if "AFFIRM" in account_name.upper():
        return "Affirm"
    return "Morgan Stanley Client Serv"


def build_external_id(
    posted_at_iso: str,
    desc_clean: str,
    amount_cents: int,
    account_id: str,
    seq: int,
) -> str:
    """Build a deterministic external_id from transaction fields.

    Includes a sequence counter to disambiguate same-day, same-amount,
    same-description transactions (e.g. two identical Zelle payments).
    """
    hash_input = f"{posted_at_iso}|{desc_clean}|{amount_cents}|{account_id}|seq{seq}"
    digest = hashlib.sha256(hash_input.encode()).hexdigest()[:24]
    return f"xlsx_{digest}"


def parse_xlsx(
    xlsx_path: str,
    *,
    skip_patterns: set[str],
    item_id: str,
) -> tuple[list[dict[str, object]], dict[str, int], dict[str, str]]:
    """Parse the XLSX file and return normalized transaction dicts.

    Args:
        xlsx_path: Path to the Activity Report XLSX file.
        skip_patterns: Account ID patterns to skip.
        item_id: Plaid item_id to associate with imported transactions.

    Returns:
        Tuple of (rows, account_stats, skipped_reasons).
        - rows: List of dicts ready for bulk_upsert_plaid_transactions.
        - account_stats: Counter of account_name -> row count.
        - skipped_reasons: Pattern -> reason for each skipped pattern.
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["MoneyMovement"]

    # Validate headers
    header_row = [cell.value for cell in ws[HEADER_ROW]]
    for idx, expected in enumerate(EXPECTED_HEADERS):
        actual = header_row[idx]
        if actual != expected:
            print(
                f"WARNING: Column {idx} header mismatch: "
                f"expected {expected!r}, got {actual!r}",
                file=sys.stderr,
            )

    rows: list[dict[str, object]] = []
    seen_hashes: Counter[str] = Counter()
    account_stats: Counter[str] = Counter()
    skipped_reasons: dict[str, str] = {}

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        account_name = row[2] or ""
        if not account_name:
            continue

        # Check skip patterns
        should_skip = False
        for pattern in skip_patterns:
            if pattern in str(account_name):
                should_skip = True
                reason = PLAID_COVERED_PATTERNS.get(pattern, "user-specified skip")
                skipped_reasons[pattern] = reason
                break
        if should_skip:
            continue

        txn_date_raw = row[1]
        activity = str(row[4] or "")
        description = str(row[5] or "")
        amount_raw = row[9]

        if amount_raw is None or txn_date_raw is None:
            continue

        # Parse date
        posted_at: date
        if isinstance(txn_date_raw, str):
            posted_at = datetime.strptime(txn_date_raw, "%m/%d/%Y").date()
        elif isinstance(txn_date_raw, datetime):
            posted_at = txn_date_raw.date()
        elif isinstance(txn_date_raw, date):
            posted_at = txn_date_raw
        else:
            continue  # skip unrecognized date types

        account_id = extract_account_id(str(account_name))
        desc_clean = " ".join(description.split())
        desc_first_line = description.split("\n")[0].strip()
        merchant_descriptor = f"{activity} {desc_first_line}".strip()
        amount_cents = int(round(float(str(amount_raw)) * 100))

        # Deterministic external_id with sequence counter
        base_key = f"{posted_at.isoformat()}|{desc_clean}|{amount_cents}|{account_id}"
        seen_hashes[base_key] += 1
        seq = seen_hashes[base_key]

        external_id = build_external_id(
            posted_at.isoformat(), desc_clean, amount_cents, account_id, seq
        )

        rows.append(
            {
                "external_id": external_id,
                "source": SOURCE,
                "account_id": account_id,
                "item_id": item_id,
                "posted_at": posted_at,
                "amount_cents": amount_cents,
                "currency": "USD",
                "merchant_descriptor": merchant_descriptor,
                "institution": resolve_institution(str(account_name)),
            }
        )
        account_stats[f"{account_name} ({account_id})"] += 1

    # Verify uniqueness
    ext_ids = [r["external_id"] for r in rows]
    dupes = [eid for eid, cnt in Counter(ext_ids).items() if cnt > 1]
    if dupes:
        raise ValueError(
            f"Duplicate external_ids detected ({len(dupes)} duplicates). "
            "Hash function needs more disambiguation fields."
        )

    return rows, dict(account_stats), skipped_reasons


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------


def import_transactions(
    rows: list[dict[str, object]],
    *,
    dry_run: bool = True,
) -> None:
    """Import parsed rows into the database.

    Args:
        rows: Normalized transaction dicts from parse_xlsx.
        dry_run: If True, only print what would happen.
    """
    from transactoid.adapters.db.facade import DB

    db_url = os.environ.get("DATABASE_URL") or "sqlite:///:memory:"
    db = DB(db_url)

    if dry_run:
        print("\n[DRY RUN] Would upsert plaid_transactions and insert derived.")
        print("[DRY RUN] Re-run with --apply to execute.\n")
        return

    from sqlalchemy import text

    # Step 1: Upsert plaid_transactions
    print(f"\nUpserting {len(rows)} plaid_transactions...")
    plaid_ids: list[int] = db.bulk_upsert_plaid_transactions(rows)
    print(f"  Got {len(plaid_ids)} plaid_transaction_ids")

    # Step 2: Check which already have derived rows
    with db.session() as session:
        result = session.execute(
            text(
                "SELECT plaid_transaction_id FROM derived_transactions "
                "WHERE plaid_transaction_id = ANY(:ids)"
            ),
            {"ids": plaid_ids},
        )
        existing_derived_pids: set[int] = {r[0] for r in result.fetchall()}

    # Step 3: Build derived data for new rows only
    derived_data: list[dict[str, object]] = []
    for row, plaid_id in zip(rows, plaid_ids, strict=True):
        if plaid_id in existing_derived_pids:
            continue
        derived_data.append(
            {
                "plaid_transaction_id": plaid_id,
                "external_id": row["external_id"],
                "amount_cents": row["amount_cents"],
                "posted_at": row["posted_at"],
                "merchant_descriptor": row["merchant_descriptor"],
                "reporting_mode": "DEFAULT_INCLUDE",
            }
        )

    skipped = len(plaid_ids) - len(derived_data)
    print(
        f"Inserting {len(derived_data)} derived_transactions "
        f"(skipping {skipped} existing)..."
    )
    if derived_data:
        derived_ids = db.bulk_insert_derived_transactions(derived_data)
        print(f"  Created {len(derived_ids)} derived_transactions")
    else:
        print("  Nothing to insert (all rows already exist)")

    # Step 4: Verify
    print("\n--- Verification ---")
    with db.session() as session:
        r = session.execute(
            text(
                "SELECT COUNT(*), MIN(posted_at), MAX(posted_at) "
                "FROM plaid_transactions WHERE source = :src"
            ),
            {"src": SOURCE},
        ).fetchone()
        if r is None:
            print("No plaid_transactions found", file=sys.stderr)
            return
        print(f"plaid_transactions (source={SOURCE}): count={r[0]}, {r[1]} to {r[2]}")

        r = session.execute(
            text(
                "SELECT COUNT(*), MIN(dt.posted_at), MAX(dt.posted_at) "
                "FROM derived_transactions dt "
                "JOIN plaid_transactions pt "
                "  ON dt.plaid_transaction_id = pt.plaid_transaction_id "
                "WHERE pt.source = :src"
            ),
            {"src": SOURCE},
        ).fetchone()
        if r is None:
            print("No derived_transactions found", file=sys.stderr)
            return
        print(f"derived_transactions (source={SOURCE}): count={r[0]}, {r[1]} to {r[2]}")

        # Per-account breakdown
        print("\n--- Per-account breakdown ---")
        acct_rows = session.execute(
            text(
                "SELECT pt.account_id, COUNT(*), MIN(dt.posted_at), MAX(dt.posted_at) "
                "FROM derived_transactions dt "
                "JOIN plaid_transactions pt "
                "  ON dt.plaid_transaction_id = pt.plaid_transaction_id "
                "WHERE pt.source = :src "
                "GROUP BY pt.account_id ORDER BY COUNT(*) DESC"
            ),
            {"src": SOURCE},
        ).fetchall()
        for r in acct_rows:
            print(f"  {r[0]:8s} | {r[1]:4d} txns | {r[2]} to {r[3]}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import Morgan Stanley XLSX transactions into transactoid.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--xlsx",
        required=True,
        help="Path to the Activity Report XLSX file",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        default=False,
        help="Actually write to the database (default: dry run)",
    )
    parser.add_argument(
        "--include-plaid-covered",
        action="store_true",
        default=False,
        help="Include accounts that are already covered by Plaid sync",
    )
    parser.add_argument(
        "--item-id",
        default=DEFAULT_ITEM_ID,
        help=(
            "Plaid item_id to associate with imported rows"
            f" (default: {DEFAULT_ITEM_ID})"
        ),
    )

    args = parser.parse_args()

    # Resolve XLSX path
    xlsx_path = os.path.expanduser(args.xlsx)
    if not os.path.isfile(xlsx_path):
        print(f"Error: File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    # Determine skip patterns
    skip_patterns: set[str] = set()
    if not args.include_plaid_covered:
        skip_patterns = set(PLAID_COVERED_PATTERNS.keys())

    # Load env for DB connection
    env_path = Path(".") / ".env"
    load_dotenv(dotenv_path=env_path, override=False)

    # Parse
    print(f"Reading: {xlsx_path}")
    rows, account_stats, skipped_reasons = parse_xlsx(
        xlsx_path, skip_patterns=skip_patterns, item_id=args.item_id
    )

    print(f"\nParsed {len(rows)} transactions across {len(account_stats)} accounts:")
    for acct, cnt in sorted(account_stats.items(), key=lambda x: -x[1]):
        print(f"  {cnt:4d} | {acct}")

    if skipped_reasons:
        print(f"\nSkipped accounts ({len(skipped_reasons)}):")
        for pattern, reason in skipped_reasons.items():
            print(f"  {pattern}: {reason}")

    if rows:
        dates = [r["posted_at"] for r in rows]
        print(f"\nDate range: {min(dates)} to {max(dates)}")  # type: ignore[type-var]

    # Import
    if not rows:
        print("\nNothing to import.")
        return

    import_transactions(rows, dry_run=not args.apply)

    if not args.apply:
        print("Re-run with --apply to execute the import.")


if __name__ == "__main__":
    main()
